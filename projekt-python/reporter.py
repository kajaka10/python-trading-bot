"""
reporter.py - Dzienny raport P&L w Excelu
"""

import os
import re
import json
import logging
from datetime import datetime, date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import trade_logger

BASE_DIR   = Path(__file__).parent
LOG_FILE   = BASE_DIR / "bot.log"
REPORT_DIR = BASE_DIR / "raporty"
REPORT_DIR.mkdir(exist_ok=True)

C_HEADER_BG  = "1F3864"
C_HEADER_FG  = "FFFFFF"
C_PROFIT     = "E2EFDA"
C_LOSS       = "FCE4D6"
C_SUMMARY_BG = "D6E4F0"
C_TITLE_BG   = "2E75B6"
C_ALT_ROW    = "F5F5F5"

THIN   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_log(target_date: date) -> list:
    if not LOG_FILE.exists():
        return []
    date_str = target_date.strftime("%Y-%m-%d")

    pattern_trade = re.compile(
        r"(\d{4}-\d{2}-\d{2} \d{2}:\d{2}).*?"
        r"([A-Z_]+): (BUY|SELL) \| SL=([\d.]+) TP=([\d.]+) \| Pewnosc: (LOW|MEDIUM|HIGH) \| Powod: (.+)"
    )
    pattern_close = re.compile(
        r"(\d{4}-\d{2}-\d{2} \d{2}:\d{2}).*?"
        r"([A-Z0-9_]+): Pozycja .+ \((BUY|SELL)\) zamknięta przez SL/TP, P&L=([-\d.]+)"
    )

    events = []
    closed_pnl = {}  # symbol -> pnl z zamknięcia

    with open(LOG_FILE, encoding="utf-8") as f:
        for line in f:
            if date_str not in line:
                continue
            if "INFO" not in line:
                continue

            mc = pattern_close.search(line)
            if mc:
                sym = mc.group(2)
                pnl = float(mc.group(4))
                closed_pnl[sym] = closed_pnl.get(sym, 0) + pnl
                events.append({
                    "time":   mc.group(1)[11:16],
                    "symbol": sym,
                    "action": mc.group(3),
                    "sl":     None,
                    "tp":     None,
                    "reason": "SL hit" if pnl <= 0 else "TP hit",
                    "status": "CLOSED",
                    "pnl":    pnl,
                })
                continue

            m = pattern_trade.search(line)
            if m:
                events.append({
                    "time":   m.group(1)[11:16],
                    "symbol": m.group(2),
                    "action": m.group(3),
                    "sl":     float(m.group(4)),
                    "tp":     float(m.group(5)),
                    "reason": m.group(7).strip()[:80],
                    "status": "OK",
                    "pnl":    None,
                })

    return events


def _hdr(cell, text, bg=C_HEADER_BG, fg=C_HEADER_FG):
    cell.value = text
    cell.font  = Font(name="Arial", bold=True, size=11, color=fg)
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def _cell(cell, value, bold=False, color="000000", align="left",
          num_fmt=None, bg=None):
    cell.value = value
    cell.font  = Font(name="Arial", bold=bold, size=10, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = BORDER
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        cell.number_format = num_fmt


def build_trade_log(ws, events):
    ws.title = "Dziennik"
    cols   = ["Czas", "Symbol", "Akcja", "Stop Loss", "Take Profit",
              "Uzasadnienie AI", "Status", "P&L (USD)"]
    widths = [8, 12, 8, 12, 12, 45, 10, 12]
    for i, (h, w) in enumerate(zip(cols, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        _hdr(ws.cell(1, i), h)
    ws.row_dimensions[1].height = 28

    for r, ev in enumerate(events, start=2):
        alt = C_ALT_ROW if r % 2 == 0 else "FFFFFF"
        action = ev["action"]
        if action == "BUY":
            act_color, act_bg = "1F6B2E", "E2EFDA"
        elif action == "SELL":
            act_color, act_bg = "9C2A2A", "FCE4D6"
        else:
            act_color, act_bg = "7F6000", "FFF2CC"

        _cell(ws.cell(r, 1), ev["time"],   align="center", bg=alt)
        _cell(ws.cell(r, 2), ev["symbol"], bold=True, align="center", bg=alt)
        _cell(ws.cell(r, 3), action, bold=True,
              color=act_color, align="center", bg=act_bg)
        _cell(ws.cell(r, 4),
              ev["sl"] if ev["sl"] is not None else "—",
              align="right", bg=alt)
        _cell(ws.cell(r, 5),
              ev["tp"] if ev["tp"] is not None else "—",
              align="right", bg=alt)
        _cell(ws.cell(r, 6), ev["reason"], bg=alt)
        _cell(ws.cell(r, 7), ev["status"], align="center", bg=alt)
        pnl = ev.get("pnl")
        if pnl is None:
            _cell(ws.cell(r, 8), "—", align="center", bg=alt)
        else:
            pnl_bg  = C_PROFIT if pnl >= 0 else "FCE4D6"
            pnl_col = "1F6B2E" if pnl >= 0 else "9C2A2A"
            _cell(ws.cell(r, 8), pnl, bold=True, color=pnl_col,
                  align="right", num_fmt="#,##0.00", bg=pnl_bg)

    ws.freeze_panes = "A2"


def build_summary(ws, events, target_date):
    ws.title = "Podsumowanie"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = f"Raport dzienny — {target_date.strftime('%d.%m.%Y')}"
    t.font  = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.fill  = PatternFill("solid", fgColor=C_TITLE_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    buy_count   = sum(1 for e in events if e["action"] == "BUY"  and e["status"] == "OK")
    sell_count  = sum(1 for e in events if e["action"] == "SELL" and e["status"] == "OK")
    wait_count  = sum(1 for e in events if e["status"] == "WAIT")
    closed      = [e for e in events if e["status"] == "CLOSED"]
    wins        = sum(1 for e in closed if (e["pnl"] or 0) > 0)
    losses      = sum(1 for e in closed if (e["pnl"] or 0) <= 0)
    total_pnl   = sum(e["pnl"] or 0 for e in closed)
    win_rate    = (wins / len(closed) * 100) if closed else 0

    # Dane spreadow z trade_logger
    date_str = target_date.isoformat()
    spread_trades = trade_logger.get_trades_for_date(date_str)
    spread_total = sum(t.get("spread_cost", 0) or 0 for t in spread_trades)
    theor_total  = sum(t.get("theoretical_pnl_no_spread", 0) or 0 for t in spread_trades)
    avg_spread   = (spread_total / len(spread_trades)) if spread_trades else 0

    rows = [
        ("Sygnaly BUY",                       buy_count),
        ("Sygnaly SELL",                      sell_count),
        ("Sygnaly WAIT",                      wait_count),
        ("Zamkniete pozycje",                 len(closed)),
        ("Zyskowne",                          wins),
        ("Stratne",                           losses),
        ("Win rate",                          f"{win_rate:.1f}%"),
        ("Laczny P&L",                        f"${total_pnl:.2f}"),
        ("Koszt spreadow",                    f"${spread_total:.2f}"),
        ("Theoretical P&L (bez spreadow)",    f"${theor_total:.2f}"),
        ("Sredni spread / trade",             f"${avg_spread:.2f}"),
    ]

    for i, (label, value) in enumerate(rows, start=2):
        ws.row_dimensions[i].height = 22
        _cell(ws.cell(i, 1), label, bg=C_SUMMARY_BG)
        _cell(ws.cell(i, 2), value, align="right", bg=C_SUMMARY_BG)



def build_spreads(ws, target_date):
    """Zakladka 'Spready' - dane z trade_logger.json (po dealId)."""
    ws.title = "Spready"
    date_str = target_date.isoformat()
    trades = trade_logger.get_trades_for_date(date_str)

    # Naglowek tabeli transakcji
    cols = ["Czas otw.", "Symbol", "Kier.", "Size",
            "Open Bid", "Open Ask", "Spread", "Koszt spreadu ($)",
            "Czas zamk.", "Gross P&L", "Theoretical P&L"]
    widths = [10, 12, 8, 8, 12, 12, 10, 16, 10, 12, 16]
    for i, (h, w) in enumerate(zip(cols, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        _hdr(ws.cell(1, i), h)
    ws.row_dimensions[1].height = 28

    # Sortuj po czasie otwarcia
    trades_sorted = sorted(trades, key=lambda t: t.get("open_time", ""))

    for r, t in enumerate(trades_sorted, start=2):
        alt = C_ALT_ROW if r % 2 == 0 else "FFFFFF"
        direction = t.get("direction", "")
        if direction == "BUY":
            d_col, d_bg = "1F6B2E", "E2EFDA"
        else:
            d_col, d_bg = "9C2A2A", "FCE4D6"

        open_t  = t.get("open_time", "")[11:16]
        close_t = t.get("close_time", "")[11:16]
        gross   = t.get("gross_pnl", 0) or 0
        theor   = t.get("theoretical_pnl_no_spread", 0) or 0
        pnl_bg  = C_PROFIT if gross >= 0 else "FCE4D6"
        pnl_col = "1F6B2E" if gross >= 0 else "9C2A2A"
        th_bg   = C_PROFIT if theor >= 0 else "FCE4D6"
        th_col  = "1F6B2E" if theor >= 0 else "9C2A2A"

        _cell(ws.cell(r, 1),  open_t, align="center", bg=alt)
        _cell(ws.cell(r, 2),  t.get("symbol", ""), bold=True, align="center", bg=alt)
        _cell(ws.cell(r, 3),  direction, bold=True, color=d_col, align="center", bg=d_bg)
        _cell(ws.cell(r, 4),  t.get("size", 0), align="right", num_fmt="0.00", bg=alt)
        _cell(ws.cell(r, 5),  t.get("open_bid", 0), align="right", num_fmt="#,##0.00", bg=alt)
        _cell(ws.cell(r, 6),  t.get("open_ask", 0), align="right", num_fmt="#,##0.00", bg=alt)
        _cell(ws.cell(r, 7),  t.get("spread_open", 0), align="right", num_fmt="0.0000", bg=alt)
        _cell(ws.cell(r, 8),  t.get("spread_cost", 0), align="right", num_fmt="$#,##0.00", bg=alt)
        _cell(ws.cell(r, 9),  close_t, align="center", bg=alt)
        _cell(ws.cell(r, 10), gross, bold=True, color=pnl_col, align="right", num_fmt="$#,##0.00", bg=pnl_bg)
        _cell(ws.cell(r, 11), theor, bold=True, color=th_col,  align="right", num_fmt="$#,##0.00", bg=th_bg)

    # ====== Sekcja Per Symbol (po pustym wierszu) ======
    start_row = len(trades_sorted) + 4
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
    title = ws.cell(start_row, 1)
    title.value = "Per Symbol"
    title.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=C_TITLE_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row].height = 24

    sym_cols = ["Symbol", "Trades", "Wins", "Win %",
                "Suma spreadow ($)", "Gross P&L", "Theoretical P&L", "% zjedzone przez spread"]
    hdr_row = start_row + 1
    for i, h in enumerate(sym_cols, 1):
        _hdr(ws.cell(hdr_row, i), h)

    # Agreguj per symbol
    per_sym = {}
    for t in trades_sorted:
        s = t.get("symbol", "")
        d = per_sym.setdefault(s, {"trades": 0, "wins": 0, "spread_sum": 0,
                                    "gross": 0, "theor": 0})
        d["trades"]     += 1
        d["wins"]       += 1 if (t.get("gross_pnl", 0) or 0) > 0 else 0
        d["spread_sum"] += t.get("spread_cost", 0) or 0
        d["gross"]      += t.get("gross_pnl", 0) or 0
        d["theor"]      += t.get("theoretical_pnl_no_spread", 0) or 0

    for j, (sym, d) in enumerate(sorted(per_sym.items()), start=hdr_row + 1):
        alt = C_ALT_ROW if j % 2 == 0 else "FFFFFF"
        win_pct = (d["wins"] / d["trades"] * 100) if d["trades"] else 0
        # % zjedzone: ile spread to wzgledem theoretical (gdy theor != 0)
        if abs(d["theor"]) > 0.01:
            eaten_pct = abs(d["spread_sum"] / d["theor"]) * 100
            eaten_str = f"{eaten_pct:.1f}%"
        else:
            eaten_str = "—"
        _cell(ws.cell(j, 1), sym, bold=True, align="center", bg=alt)
        _cell(ws.cell(j, 2), d["trades"], align="center", bg=alt)
        _cell(ws.cell(j, 3), d["wins"], align="center", bg=alt)
        _cell(ws.cell(j, 4), f"{win_pct:.1f}%", align="center", bg=alt)
        _cell(ws.cell(j, 5), d["spread_sum"], align="right", num_fmt="$#,##0.00", bg=alt)

        gross_col = "1F6B2E" if d["gross"] >= 0 else "9C2A2A"
        theor_col = "1F6B2E" if d["theor"] >= 0 else "9C2A2A"
        _cell(ws.cell(j, 6), d["gross"], bold=True, color=gross_col,
              align="right", num_fmt="$#,##0.00", bg=alt)
        _cell(ws.cell(j, 7), d["theor"], bold=True, color=theor_col,
              align="right", num_fmt="$#,##0.00", bg=alt)
        _cell(ws.cell(j, 8), eaten_str, align="center", bg=alt)

    ws.freeze_panes = "A2"


def generate_report(target_date: date = None) -> Path:
    if target_date is None:
        target_date = date.today()

    events = parse_log(target_date)
    logging.info(f"REPORTER: Znaleziono {len(events)} zdarzen dla {target_date}")

    wb = Workbook()
    wb.remove(wb.active)
    build_trade_log(wb.create_sheet(), events)
    build_summary(wb.create_sheet(), events, target_date)
    build_spreads(wb.create_sheet(), target_date)

    fname = f"raport_{target_date.isoformat()}.xlsx"
    fpath = REPORT_DIR / fname
    wb.save(fpath)

    logging.info(f"REPORTER: Zapisano raport -> {fpath}")
    print(f"Raport wygenerowany: {fpath} ({len(events)} zdarzen)")
    return fpath


if __name__ == "__main__":
    import sys
    d = date.fromisoformat(sys.argv[1]) if len(sys.argv) > 1 else date.today()
    generate_report(d)
